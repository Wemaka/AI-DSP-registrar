"""
utils/excel_export.py
Экспорт журнала в xlsx с форматированием (Требование 5 — таблицы).
"""

from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CLASS_COLORS = {
    "Штатная":   "C6EFCE",
    "Тех. сбой": "FFEB9C",
    "АВАРИЯ":    "FFC7CE",
}

COL_WIDTHS = {
    "Время": 20, "Поезд": 10, "Путь": 8, "Статус": 30,
    "Класс события": 14, "Уверенность ML": 14, "Примечание": 35, "Источник": 10,
}


def export_to_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал движения"

    # Заголовок листа
    n_cols = len(df.columns)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws["A1"] = f"Журнал движения поездов — {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws["A1"].fill = PatternFill("solid", start_color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Шапка таблицы
    header_fill = PatternFill("solid", start_color="2E75B6")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Данные строки
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 3
        event_class = str(row.get("Класс события", ""))
        fill_color = CLASS_COLORS.get(event_class, "FFFFFF")

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill = PatternFill("solid", start_color=fill_color)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", size=10)

    # Ширина колонок
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)

    # Итоговая строка
    last_data_row = len(df) + 2
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=1, value="Итого:").font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=2, value=f"=COUNTA(B3:B{last_data_row})").font = Font(bold=True, name="Arial")

    # Легенда
    legend_row = total_row + 2
    ws.cell(row=legend_row, column=1, value="Легенда:").font = Font(bold=True, name="Arial")
    for i, (cls, color) in enumerate(CLASS_COLORS.items(), start=1):
        cell = ws.cell(row=legend_row + i, column=1, value=cls)
        cell.fill = PatternFill("solid", start_color=color)
        cell.font = Font(name="Arial", size=9)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
